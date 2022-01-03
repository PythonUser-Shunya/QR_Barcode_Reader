# 実行は別ファイルから
# pyinstaller QR_Barcode.py --onefile --noconsole --add-binary "C:\Users\***\anaconda3\envs\barcode\Lib\site-packages\pyzbar\*.dll;pyzbar"
from pyzbar.pyzbar import decode
import cv2
import re
from collections import Counter
import numpy as np
import warnings
import openpyxl

# 警告を非表示にする
warnings.simplefilter("ignore")
# 読み込んだ情報を格納する
number_list = []
font = cv2.FONT_HERSHEY_SIMPLEX

def save_battery(read_list):
    """
    電池用。読み込んだ情報をcsvファイルに保存する。

    Parameters
    ----------
    read_list : list
        読み込んだ情報
    """
    np.savetxt("battery_list.csv", read_list, delimiter=",", fmt="%s")


def save_SD(read_list):
    """
    SD用。読み込んだ情報をcsvファイルに保存する。

    Parameters
    ----------
    read_list : list
        読み込んだ情報
    """
    np.savetxt("SD_list.csv", read_list, delimiter=",", fmt="%s")


def save_camera(read_list):
    """
    カメラ用。読み込んだ情報をcsvファイルに保存する。

    Parameters
    ----------
    read_list : list
        読み込んだ情報
    """
    np.savetxt("camera_list.csv", read_list, delimiter=",", fmt="%s")


def load_battery_list():
    """
    電池用。ファイルの情報（履歴）を読み込んでリスト型にする。
    Returns
    -------
    load_list : list
        読み込んだ情報
    """
    file = "battery_list.csv"
    load_list = list(np.loadtxt(
        file, delimiter=",", dtype="unicode", ndmin=1))
    load_list = list(set(load_list))
    return load_list


def load_SD_list():
    """
    SD用。ファイルの情報（履歴）を読み込んでリスト型にする。
    Returns
    -------
    load_list : list
        読み込んだ情報
    """
    file = "SD_list.csv"
    load_list = list(np.loadtxt(
        file, delimiter=",", dtype="unicode", ndmin=1))
    load_list = list(set(load_list))
    return load_list


def load_camera_list():
    """
    カメラ用。ファイルの情報（履歴）を読み込んでリスト型にする。
    Returns
    -------
    load_list : list
        読み込んだ情報
    """
    file = "camera_list.csv"
    load_list = list(np.loadtxt(
        file, delimiter=",", dtype="unicode", ndmin=1))
    load_list = list(set(load_list))
    return load_list


def make_battery_list():
    """
    電池用。もし".csv"があるならその情報（履歴）を使う。なかったら新しく空のリストを作る

    Returns
    -------
    compiled_list : list
        読み込んだ情報
    """
    try:
        compiled_list = load_battery_list()
    except OSError:
        compiled_list = []
    return compiled_list


def make_SD_list():
    """
    SD用。もし".csv"があるならその情報（履歴）を使う。なかったら新しく空のリストを作る

    Returns
    -------
    compiled_list : list
        読み込んだ情報
    """
    try:
        compiled_list = load_SD_list()
    except OSError:
        compiled_list = []
    return compiled_list


def make_camera_list():
    """
    カメラ用。もし".csv"があるならその情報（履歴）を使う。なかったら新しく空のリストを作る

    Returns
    -------
    compiled_list : list
        読み込んだ情報
    """
    try:
        compiled_list = load_camera_list()
    except OSError:
        compiled_list = []
    return compiled_list


def make_new_excel_battery(xlsx_file_path, place, read_list):
    """
    電池用。読み取った情報を元にExcelファイルを更新する

    Parameters
    ----------
    xlsx_file_path : str
        Excelファイルのパス
    place : str
        使用場所
    read_list : list
        読み込んだ情報
    
    Returns
    -------
    True : bool
        更新が成功したらTrue
    """
    # Excelファイルを開く
    wb = openpyxl.load_workbook(xlsx_file_path)
    # 指定のシートを開く
    ws = wb.worksheets[0]
    # A列を探索。もし読み込んだ番号と一致する番号があればその行の「場所」カラムを更新する
    for cell_a in ws["A"]:
        if str(cell_a.value) in read_list:
            cell_a_row = cell_a.row
            ws[f"C{cell_a_row}"].value = place
    wb.save(xlsx_file_path)
    return True


def make_new_excel_SD(xlsx_file_path, place, read_list, user):
    """
    SD用。読み取った情報を元にExcelファイルを更新する

    Parameters
    ----------
    xlsx_file_path : str
        Excelファイルのパス
    place : str
        使用場所
    read_list : list
        読み込んだ情報

    Returns
    -------
    True : bool
        更新が成功したらTrue
    """
    # Excelファイルを開く
    wb = openpyxl.load_workbook(xlsx_file_path)
    # 指定のシートを開く
    ws = wb.worksheets[0]
    # A列を探索。もし読み込んだ番号と一致する番号があればその行の「場所」と「使用者」のカラムを更新する。
    for cell_a in ws["A"]:
        if cell_a.value in read_list:
            cell_a_row = cell_a.row
            ws[f"E{cell_a_row}"].value = place
            ws[f"F{cell_a_row}"].value = user
    wb.save(xlsx_file_path)
    return True


def make_new_excel_camera(xlsx_file_path, place, read_list, user):
    """
    カメラ用。読み取った情報を元にExcelファイルを更新する

    Parameters
    ----------
    xlsx_file_path : str
        Excelファイルのパス
    place : str
        使用場所
    read_list : list
        読み込んだ情報

    Returns
    -------
    True : bool
        更新が成功したらTrue
    """
    # Excelファイルを開く
    wb = openpyxl.load_workbook(xlsx_file_path)
    # 指定のシートを開く
    ws = wb.worksheets[0]
    # A列を探索。もし読み込んだ番号と一致する番号があればその行の「場所」と「使用者」のカラムを更新する。
    for cell_a in ws["A"]:
        if str(cell_a.value) in read_list:
            cell_a_row = cell_a.row
            ws[f"B{cell_a_row}"].value = place
            ws[f"C{cell_a_row}"].value = user
    wb.save(xlsx_file_path)
    return True


def over30_and_under4(number_list):
    """
    電池とカメラ用のフィルター。30回以上の重複かつ4桁以下かつバーコートのみ許す

    Parameters
    ----------
    number_list : list
        読み込んだ情報
    
    Returns
    -------
    number_list_counter_30 : list
        条件に一致する番号のリスト
    """
    # リストに格納されている番号のそれぞれの出現回数を数える
    number_list_counter = Counter(number_list)
    number_list_counter_30 = []
    # 出現回数が30回以上かつ4桁以下の情報のみを抽出
    for number_list_counter in number_list_counter.items():
        if number_list_counter[1] >= 30 and len(str(number_list_counter[0])) <= 4:
            number_list_counter_30.append(number_list_counter[0])
    return number_list_counter_30


def over30_and_under5_list(number_list):
    """
    SD用フィルター。30回以上の重複かつ5桁以下かつアンダーバーがある情報のみ許す

    Parameters
    ----------
    number_list : list
        読み込んだ情報
    
    Returns
    -------
    number_list_counter_30 : list
        条件に一致する番号のリスト
    """
    number_list_counter = Counter(number_list)
    number_list_counter_30 = []
    # 出現回数が30回以上かつ5文字以下かつ「_」を含む情報のみを抽出
    for number_list_counter in number_list_counter.items():
        if number_list_counter[1] >= 30 and len(number_list_counter_30[0]) <= 5 and "_" in number_list_counter_30[0]:
            number_list_counter_30 = list(number_list_counter[0])
    return number_list_counter_30

def text(frame, compiled_list):
    """
    読み込んだ番号の総数を表示させる

    Parameters
    ----------
    frame : cv2.VideoCapture
        カメラから取得したフレーム
    compiled_list : list
        フィルターにかけられた後の読み込んだ情報
    """
    # 読みやすいように太字(黒)の上に細字(白)で表示
    cv2.putText(frame, f'COUNT: {len(compiled_list)}', (0, 35),
                font, 1, (0, 0, 0), 3, cv2.LINE_AA)
    cv2.putText(frame, f'COUNT: {len(compiled_list)}', (0, 35),
                font, 1, (255, 255, 255), 1, cv2.LINE_AA)

def draw_rectangle(frame, x, y, w, h):
    """
    バーコードに矩形描画

    Parameters
    ----------
    frame : cv2.VideoCapture
        カメラから取得したフレーム
    x : int
        矩形の左上のx座標
    y : int
        矩形の左上のy座標
    w : int
        矩形の幅
    h : int
        矩形の高さ
    """
    # バーコードを囲むように矩形を描画
    cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 0, 255), 2)
    # バーコードの少し上に塗りつぶした矩形を描画。数字を見やすくするため
    cv2.rectangle(frame, (x, y-25), (x+60, y), (0, 0, 255), -1)

def read_battery(camera_number, compiled_list):
    """
    電池のバーコードを読み取る

    Parameters
    ----------
    camera_number : int
        カメラの番号
    compiled_list : list
        履歴から読み込んだ情報。これに追加していく。履歴が無かったらこれはからのリストになる
    
    Returns
    -------
    compiled_list : list
        最終的な番号のリスト
    """
    # 数字の正規表現
    re_compile = re.compile('^[0-9]+$')
    # カメラ読み込み
    cap = cv2.VideoCapture(int(camera_number))
    while cap.isOpened():
        ret, frame = cap.read()
        if ret == True:
            # フレーム内のバーコードを検出
            d = decode(frame)
            if d:
                # バーコードが検出されたらひとつずつ処理を施す
                for barcode in d:
                    if barcode.type == 'CODE128':
                        x, y, w, h = barcode.rect
                        # 矩形描画
                        draw_rectangle(frame, x, y, w, h)
                        barcodeData = barcode.data.decode('utf-8')
                        # 読み込んだ情報をバーコードの少し上に表示
                        frame = cv2.putText(
                            frame, barcodeData, (x, y-10), font, .5, (255, 255, 255), 1, cv2.LINE_AA)
                        # 読み込んだ情報をリストに追加
                        number_list.append(barcodeData)
                        # フィルターをかける
                        number_list_counter_30 = over30_and_under4(
                            number_list)
                        # さらに正規表現のフィルターをかける
                        for c in filter(re_compile.match, number_list_counter_30):
                            compiled_list.append(c)
                            # 重複は許さない
                            compiled_list = list(set(compiled_list))
            # 文字列の表示
            text(frame, compiled_list)
            cv2.imshow('frame', frame)
        key = cv2.waitKey(10)
        if key == 27:  # ESCkey
            break

    cap.release()
    cv2.destroyAllWindows()
    number_list.clear()
    return compiled_list

def read_camera(camera_number, compiled_list):
    """
    カメラのバーコードを読み取る

    Parameters
    ----------
    camera_number : int
        カメラの番号
    compiled_list : list
        履歴から読み込んだ情報。これに追加していく。履歴が無かったらこれはからのリストになる

    Returns
    -------
    compiled_list : list
        最終的な番号のリスト
    """
    re_compile = re.compile('^[0-9]+$')
    cap = cv2.VideoCapture(int(camera_number))
    while cap.isOpened():
        ret, frame = cap.read()
        if ret == True:
            d = decode(frame)
            if d:
                for barcode in d:
                    if barcode.type == 'QRCODE':
                        x, y, w, h = barcode.rect
                        draw_rectangle(frame, x, y, w, h)
                        barcodeData = barcode.data.decode('utf-8')
                        frame = cv2.putText(
                            frame, barcodeData, (x, y-10), font, .5, (255, 255, 255), 1, cv2.LINE_AA)
                        number_list.append(barcodeData)
                        number_list_counter_30 = over30_and_under4(
                            number_list)
                        for c in filter(re_compile.match, number_list_counter_30):
                            compiled_list.append(c)
                            compiled_list = list(set(compiled_list))
            # 文字列の表示
            text(frame, compiled_list)
            cv2.imshow('frame', frame)
        key = cv2.waitKey(10)
        if key == 27:  # ESCkey
            break

    cap.release()
    cv2.destroyAllWindows()
    number_list.clear()
    return compiled_list

def read_SD(camera_number, compiled_list):
    """
    SDのバーコードを読み取る

    Parameters
    ----------
    camera_number : int
        カメラの番号
    compiled_list : list
        履歴から読み込んだ情報。これに追加していく。履歴が無かったらこれはからのリストになる

    Returns
    -------
    compiled_list : list
        最終的な番号のリスト
    """
    cap = cv2.VideoCapture(int(camera_number))
    while cap.isOpened():
        ret, frame = cap.read()
        if ret == True:
            d = decode(frame)
            if d:
                for barcode in d:
                    if barcode.type == 'QRCODE':
                        x, y, w, h = barcode.rect
                        draw_rectangle(frame, x, y, w, h)
                        barcodeData = barcode.data.decode('utf-8')
                        frame = cv2.putText(
                            frame, barcodeData, (x, y-10), font, .5, (255, 255, 255), 1, cv2.LINE_AA)
                        number_list.append(barcodeData)
                        number_list_counter_30 = over30_and_under5_list(
                            number_list)
                        compiled_list += (number_list_counter_30)
                        compiled_list = list(set(compiled_list))
            # 文字列の表示
            text(frame, compiled_list)
            cv2.imshow('frame', frame)
        key = cv2.waitKey(10)
        if key == 27:  # ESCkey
            break

    cap.release()
    cv2.destroyAllWindows()
    number_list.clear()
    return compiled_list
